import requests
import xlsxwriter
from bs4 import BeautifulSoup
from openpyxl import Workbook
import numpy as np
import re

# workbook = xlsxwriter.Workbook('e:\data.xlsx')
# worksheet = workbook.add_worksheet()
# title = [U'题目',U'作者',U'单位']
# worksheet.write_row('A1',title)
#workbook.close()

wb = Workbook()
ws = wb.create_sheet("data")
title = []
name = []
company = []
type = []

headers = {
    #'Accept-Encoding': 'gzip, deflate',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
    'cookies': 'session=eyJpdiI6IjZZZmZnQUZPeFlDZWhkdkRDMjR6aEE9PSIsInZhbHVlIjoib3RKQlpXV0ppZ1E1MTl6Z0tuaUdWYWc1ZzlTN1hSWmxlVHcyMUZEbjBaTVNJcHYyU1BacWFzeVM1bzRvdmZndiIsIm1hYyI6IjFhNDUyZDg2MDE0MjMyMWY5MWQyZDBkNjY1NjllODZhNjMzNTAzMjI0MDg0OGYwOGNhODQ4NGQ4ZGI3NDgwYjMifQ%3D%3D; expires=Sat, 20-Jun-2020 09:49:16 GMT; Max-Age=7200; path=/; httponly'
}
data = {
    'phone':'15797623360',
    'password':'scygj0908',
}
url = 'http://fund.sciencenet.cn/login'
sub_url = 'http://fund.sciencenet.cn/search/smallSubject?subject=H18&yearStart=2015&yearEnd=2017&filter%5Bcategory%5D%5B%5D=%E9%87%8D%E7%82%B9%E9%A1%B9%E7%9B%AE&filter%5Bcategory%5D%5B%5D=%E9%9D%A2%E4%B8%8A%E9%A1%B9%E7%9B%AE&submit=list'
sub_url_start_2015 = 'http://fund.sciencenet.cn/search/smallSubject?subject=H18&yearStart=2015&yearEnd=2015&filter%5Bcategory%5D%5B0%5D=%E9%87%8D%E7%82%B9%E9%A1%B9%E7%9B%AE&filter%5Bcategory%5D%5B1%5D=%E9%9D%A2%E4%B8%8A%E9%A1%B9%E7%9B%AE&submit=list&page='
sub_url_start_2016 = 'http://fund.sciencenet.cn/search/smallSubject?subject=H18&yearStart=2016&yearEnd=2016&filter%5Bcategory%5D%5B0%5D=%E9%87%8D%E7%82%B9%E9%A1%B9%E7%9B%AE&filter%5Bcategory%5D%5B1%5D=%E9%9D%A2%E4%B8%8A%E9%A1%B9%E7%9B%AE&submit=list&page='
sub_url_start_2017 = 'http://fund.sciencenet.cn/search/smallSubject?subject=H18&yearStart=2017&yearEnd=2017&filter%5Bcategory%5D%5B0%5D=%E9%87%8D%E7%82%B9%E9%A1%B9%E7%9B%AE&filter%5Bcategory%5D%5B1%5D=%E9%9D%A2%E4%B8%8A%E9%A1%B9%E7%9B%AE&submit=list&page='

session = requests.Session()
res = session.post(url,headers = headers,data = data)
#print(res.text)
for i in range(25):
    sub_url = sub_url_start_2017 + str(i)
    print(sub_url)
    response = session.get(sub_url, headers=headers)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')
#print(soup)
    sub_titles = soup.find_all("p", {"class": "t"})
    for sub_title in sub_titles:
        title.append(sub_title.find('a').string)
        #print(sub_url.find('a').string)
    #print(title)

    sub_names = soup.find_all("p", {"class": "ico"})
    for sub_name in sub_names:
        name.append(sub_name.find('span').find('i').string)
    #print(sub_name.find('span').find('i').string)
    #print(name)

    sub_companys = soup.find_all("p", {"class": "ico"})
    for sub_company in sub_companys:
        sub_com = sub_company.find_all('span')
        company.append(sub_com[1].find('i').string)
    #print(sub_com[1].find('i').string)
    #print(company)

    sub_types = soup.find_all("p", {"class": "ico"})
    for sub_type in sub_types:
        sub_ty = sub_type.find_all("i")
        type.append(sub_ty[2].string)
        #print(sub_ty[2].string)



ws.append(title)
ws.append(name)
ws.append(company)
ws.append(type)

wb.save("e:\data2017.xlsx")
#write_data = np.array(write_data)


# res = requests.get(url)
# print(res.text)
# cook = ""
# print(res.cookies)
# for c in res.cookies:
#     cook += c.name + "="+c.value + ";"
# print(cook)

# html = BeautifulSoup(res.text, 'html.parser')
# token = html.find_all(type='hidden')[1]['value']
# print(token)